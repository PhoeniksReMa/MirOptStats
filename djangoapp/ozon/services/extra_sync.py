from __future__ import annotations

import io
import math
from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Dict, List, Optional, Tuple

import openpyxl
import requests

from ozon.services.ozon_client import OzonClient
from ozon.services.reporting import ensure_columns, get_or_create_report, upsert_rows
from ozon.models import OzonClusterSlot
from shops.models import Shop

TURNOVER_GRADE_MAP = {
    "TURNOVER_GRADE_NONE": "нет статуса ликвидности.",
    "DEFICIT": "Хватит до 28 дней",
    "POPULAR": "Хватит на 28–56 дней",
    "ACTUAL": "Хватит на 56–120 дней",
    "SURPLUS": "Продаётся медленно, хватит > 120 дней",
    "NO_SALES": "Без продаж последние 28 дней",
    "WAS_NO_SALES": "Без продаж и остатков последние 28 дней",
    "RESTRICTED_NO_SALES": "Запрет FBO",
    "COLLECTING_DATA": "Сбор данных",
    "WAITING_FOR_SUPPLY": "Сделайте поставку для сбора данных",
    "WAS_DEFICIT": "Был дефицитным последние 56 дней",
    "WAS_POPULAR": "Был очень популярным последние 56 дней",
    "WAS_ACTUAL": "Был популярным последние 56 дней",
    "WAS_SURPLUS": "Был избыточным последние 56 дней",
}

COLOR_INDEX_MAP = {
    "WITHOUT_INDEX": "НЕТ",
    "GREEN": "ХОРОШИЙ",
    "YELLOW": "СРЕДНИЙ",
    "RED": "ПЛОХОЙ",
}

EXCLUDED_ACTIONS = {
    "Рассрочка 0-0-6 на всё РФ товары",
    "WOW-БЭК_Кэшбэк на покупку Ozon Fashion списание 2.0",
    "ВАУ баллы 50% 9 волна 3-я волна (основная)",
    "Ozon Fashion + Jardin 500 вау баллов  списание",
    "Ozon Fashion + Jardin 1000 списание",
    "[Ozon Fashion + Jardin 10 000 списание",
    "Ozon Fashion + Jardin списание 1 млн",
    "Ozon Fashion + Jardin списание 1 млн вторая",
    "Ozon Fashion + Jardin_Запасная акция 500 вау баллов списание",
    "Ozon Fashion + Jardin_Запасная акция 1000 вау баллов списание",
    "Ozon Fashion + Jardin_Запасная акция 10 000 вау баллов списание",
    "Ozon x Ростикс / Вкусная игра Номинал ВАУ-баллов - 200, количество - 25 000 шт. списание",
    "Ozon x Ростикс / Вкусная игра Номинал ВАУ-баллов - 1000, количество - 15 500 шт. списание",
    "Ozon x Ростикс / Вкусная игра Номинал ВАУ-баллов - 100 000, количество - 10 шт. списание",
    "Ozon x Ростикс / Вкусная игра Номинал ВАУ-баллов - 1 000 000, количество - 1 шт. списание",
    "Ozon Fashion + Jardin 10 000 списание",
    "Промокоды для интеграции в НГ акцию от t2 // ВАУ-баллы списание 200",
    "Промокоды для интеграции в НГ акцию от t2 // ВАУ-баллы списание 500",
    "Промокоды для интеграции в НГ акцию от t2 // ВАУ-баллы списание 500",
    "Промокоды для интеграции в НГ акцию от t2 // ВАУ-баллы списание 200",
    "Промокоды для интеграции в НГ акцию от t2 // ВАУ-баллы списание 500",
    "РК. Честная рассрочка 0-0-6",
    "РК. Честная рассрочка 0-0-12",
    "Товары со скидкой на платном хранении",
    "Рассрочка Беларусь для теста на 5% клиентов. Хайлайт Людвига",
    "РК.Рассрочка 0-0-12 до 31.01.2026",
    "Скидка 10% для сотрудников НГ товары + Книги",
    "РК.Рассрочка 0-0-6 до 31.01.3031",
    "Дополнительные промокоды для интеграции в НГ акцию от t2 // ВАУ-баллы 200 списание",
    "Дополнительные промокоды для интеграции в НГ акцию от t2 // ВАУ-баллы 500 списание",
    "Дополнительные промокоды для интеграции в НГ акцию от t2 // ВАУ-баллы 500 списание",
    "Промокоды для интеграции в акцию \"Обмен минут и ГБ\" от t2 // ВАУ-баллы (бюджет коммерции) списание 100 баллов",
    "Промокоды для интеграции в акцию \"Обмен минут и ГБ\" от t2 // ВАУ-баллы (бюджет коммерции)  200 ВАУ-баллов х 100 000 шт списание",
    "Промокоды для интеграции в акцию \"Обмен минут и ГБ\" от t2 // ВАУ-баллы (бюджет коммерции)  500 ВАУ-баллов х 50 000 шт",
}

WAREHOUSE_NAME_REPLACEMENTS = {
    "РОСТОВ_НА_ДОНУ_2": "Ростов-на-Дону",
}

RU_LOWER_WORDS = {
    "и", "в", "во", "на", "к", "ко", "о", "об", "от", "до", "за", "из", "с", "со",
    "у", "по", "при", "для", "над", "под", "без", "про",
}


def _round_two_half_up(value: Optional[object]) -> Optional[str]:
    if value is None:
        return None
    try:
        s = str(value).replace(",", ".").strip()
        d = Decimal(s)
        q = d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return format(q, "f")
    except (InvalidOperation, ValueError):
        return None


def _normalize_turnover(value: Optional[str]) -> str:
    if not value:
        return ""
    return TURNOVER_GRADE_MAP.get(value, value)


def _clean_wh_name(name: str) -> str:
    name = (name or "").strip()
    suffixes = ["_РФЦ", " РФЦ", "-РФЦ", "- РФЦ"]
    for suf in suffixes:
        if name.endswith(suf):
            return name[: -len(suf)].strip()
    return name


def _cap_ru_word(word: str, force_capital: bool = False) -> str:
    w = (word or "").strip()
    if not w:
        return w
    if w.isdigit():
        return w
    if w.isupper() and len(w) <= 3:
        return w
    lw = w.lower()
    if (not force_capital) and (lw in RU_LOWER_WORDS):
        return lw
    return lw[:1].upper() + lw[1:]


def _smart_title_ru(text: str) -> str:
    s = (text or "").strip()
    if not s:
        return s
    s = " ".join(s.replace("_", " ").split())
    words = s.split(" ")
    out_words = []
    for wi, word in enumerate(words):
        parts = word.split("-")
        out_parts = []
        for pi, part in enumerate(parts):
            force = (wi == 0 and pi == 0)
            out_parts.append(_cap_ru_word(part, force_capital=force))
        out_words.append("-".join(out_parts))
    return " ".join(out_words).strip()


def _normalize_display_name(name: str) -> str:
    name = _clean_wh_name(name)
    key = (name or "").strip().upper()
    if key in WAREHOUSE_NAME_REPLACEMENTS:
        return WAREHOUSE_NAME_REPLACEMENTS[key].strip()
    return _smart_title_ru(name).strip()


def sync_clusters(shop: Shop) -> None:
    report = get_or_create_report(shop, "clusters", "Кластеры", "Список кластеров Ozon")
    ensure_columns(report, [
        ("cluster_id", "ID", 10, "number"),
        ("name", "Название", 20, "text"),
        ("type", "Тип", 30, "text"),
    ])
    client = OzonClient(client_id=shop.client_id, api_key=shop.token)
    clusters = client.cluster_list("CLUSTER_TYPE_OZON") + client.cluster_list("CLUSTER_TYPE_CIS")
    rows: Dict[str, Dict] = {}
    for c in clusters:
        cid = c.get("cluster_id") or c.get("id")
        rows[str(cid)] = {
            "cluster_id": cid,
            "name": c.get("name") or c.get("cluster_name"),
            "type": c.get("type"),
        }
    upsert_rows(report, rows)


def sync_returns(shop: Shop) -> None:
    report = get_or_create_report(shop, "returns", "Возвраты", "Список возвратов за 30 дней")
    ensure_columns(report, [
        ("return_id", "ID возврата", 10, "text"),
        ("offer_id", "Артикул", 20, "text"),
        ("sku", "SKU", 30, "text"),
        ("quantity", "Кол-во", 40, "number"),
        ("status", "Статус", 50, "text"),
        ("reason", "Причина", 60, "text"),
        ("price", "Цена", 70, "number"),
        ("created_at", "Создан", 80, "date"),
    ])
    client = OzonClient(client_id=shop.client_id, api_key=shop.token)
    end = datetime.now(timezone.utc)
    start = end - timedelta(days=30)
    items = client.returns_list(
        start.strftime("%Y-%m-%dT%H:%M:%SZ"),
        end.strftime("%Y-%m-%dT%H:%M:%SZ"),
    )
    rows: Dict[str, Dict] = {}
    for it in items:
        rid = it.get("return_id") or it.get("id") or it.get("posting_number") or ""
        rows[str(rid)] = {
            "return_id": rid,
            "offer_id": it.get("offer_id"),
            "sku": it.get("sku"),
            "quantity": it.get("quantity"),
            "status": it.get("status"),
            "reason": it.get("reason"),
            "price": it.get("price"),
            "created_at": it.get("created_at"),
        }
    upsert_rows(report, rows)


def sync_storage(shop: Shop) -> None:
    report = get_or_create_report(shop, "storage", "Хранение", "Стоимость хранения за период")
    ensure_columns(report, [
        ("sku", "SKU", 10, "number"),
        ("cost_total", "Стоимость за период", 20, "number"),
        ("qty_paid", "Платные экземпляры", 30, "number"),
        ("forecast_28", "Прогноз 28 дней", 40, "number"),
        ("warehouses_count", "Складов", 50, "number"),
        ("note", "Склады", 60, "text"),
    ])
    client = OzonClient(client_id=shop.client_id, api_key=shop.token)
    date_to = datetime.now().date().isoformat()
    date_from = (datetime.now().date() - timedelta(days=30)).isoformat()
    code = client.report_create_placement(date_from, date_to)
    info = client.report_info(code)
    file_url = info.get("file")
    if not file_url:
        return
    resp = requests.get(file_url, timeout=60)
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    idx = {name: i for i, name in enumerate(headers)}
    # find date column
    date_col = None
    for h in headers:
        if "дата" in h.lower():
            date_col = h
            break
    required = [date_col, "Склад", "SKU", "Кол-во экземпляров", "Кол-во платных экземпляров", "Начисленная стоимость размещения"]
    if not date_col:
        return
    cost_sum_by_sku: Dict[int, float] = {}
    snap_by_sku_day: Dict[int, Dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = row[idx[date_col]]
        if not d:
            continue
        sku_raw = row[idx["SKU"]]
        try:
            sku = int(float(sku_raw))
        except Exception:
            continue
        wh = str(row[idx["Склад"]] or "").strip() or "Без названия"
        qty_total = int(row[idx["Кол-во экземпляров"]] or 0)
        qty_paid = int(row[idx["Кол-во платных экземпляров"]] or 0)
        cost = float(row[idx["Начисленная стоимость размещения"]] or 0)
        cost_sum_by_sku[sku] = cost_sum_by_sku.get(sku, 0.0) + cost
        snap_by_sku_day.setdefault(sku, {})
        snap_by_sku_day[sku].setdefault(d, {"qty_paid_total": 0, "daily_cost_total": 0.0, "wh_qty_total": {}, "warehouses_nonzero": set()})
        s = snap_by_sku_day[sku][d]
        s["qty_paid_total"] += qty_paid
        s["daily_cost_total"] += cost
        s["wh_qty_total"][wh] = s["wh_qty_total"].get(wh, 0) + qty_total
        if qty_total > 0:
            s["warehouses_nonzero"].add(wh)
    rows: Dict[str, Dict] = {}
    for sku in sorted(cost_sum_by_sku.keys()):
        cost_total = float(cost_sum_by_sku.get(sku, 0.0))
        day_map = snap_by_sku_day.get(sku, {})
        if day_map:
            available_days = sorted(day_map.keys())
            snap_day = available_days[-1]
            snap = day_map[snap_day]
            dl_qty_paid = int(snap["qty_paid_total"])
            daily_cost = float(snap["daily_cost_total"])
            forecast_28 = int(round(daily_cost * 28.0))
            wh_nonzero = sorted(set(snap["warehouses_nonzero"]))
            wh_count = len(wh_nonzero)
            note_lines = [f"🔹 {wh} - {int(snap['wh_qty_total'].get(wh, 0))} шт" for wh in wh_nonzero]
            note = "\n".join(note_lines)
        else:
            dl_qty_paid = 0
            forecast_28 = 0
            wh_count = 0
            note = ""
        rows[str(sku)] = {
            "sku": sku,
            "cost_total": int(round(cost_total)),
            "qty_paid": dl_qty_paid,
            "forecast_28": int(forecast_28),
            "warehouses_count": int(wh_count),
            "note": note,
        }
    upsert_rows(report, rows)


def sync_price_logistics(shop: Shop) -> None:
    report = get_or_create_report(shop, "price_logistics", "Цены и логистика", "Данные по ценам и логистике")
    ensure_columns(report, [
        ("product_id", "Product ID", 10, "number"),
        ("HP", "HP", 20, "number"),
        ("HQ", "HQ", 30, "number"),
        ("HR", "HR", 40, "number"),
        ("HS", "HS", 50, "number"),
        ("HT", "HT", 60, "number"),
        ("HU", "HU", 70, "number"),
        ("HV", "HV", 80, "number"),
        ("HW", "HW", 90, "number"),
        ("HX", "HX", 100, "number"),
        ("HY", "HY", 110, "number"),
        ("HZ", "HZ", 120, "number"),
        ("IA", "IA", 130, "number"),
        ("IB", "IB", 140, "number"),
        ("IC", "IC", 150, "text"),
        ("ID", "ID", 160, "text"),
        ("IE", "IE", 170, "number"),
        ("IF", "IF", 180, "number"),
        ("IG", "IG", 190, "number"),
        ("IH", "IH", 200, "number"),
        ("II", "II", 210, "number"),
        ("IJ", "IJ", 220, "text"),
        ("IK", "IK", 230, "text"),
        ("IL", "IL", 240, "number"),
        ("IM", "IM", 250, "number"),
    ])
    client = OzonClient(client_id=shop.client_id, api_key=shop.token)
    # берём product_ids из monitor
    product_ids = []
    monitor = shop.ozon_reports.filter(code="monitor").first()
    if not monitor:
        return
    for r in monitor.rows.all():
        if r.data.get("product_id"):
            product_ids.append(int(r.data["product_id"]))
    if not product_ids:
        return
    items = client.product_prices(product_ids)
    rows: Dict[str, Dict] = {}
    for item in items:
        pid = item.get("product_id")
        price_data = item.get("price") or {}
        commissions = item.get("commissions") or {}
        price_indexes = item.get("price_indexes") or {}
        marketing_actions = (item.get("marketing_actions") or {}).get("actions") or []
        acquiring = float(item.get("acquiring") or 0)
        sales_percent_fbo = float(commissions.get("sales_percent_fbo") or 0)
        sales_percent_fbs = float(commissions.get("sales_percent_fbs") or 0)
        marketing_seller_price = float(price_data.get("marketing_seller_price") or 0)
        dr_value = math.ceil((marketing_seller_price * sales_percent_fbo) / 100) if marketing_seller_price and sales_percent_fbo else 0
        ds_value = math.ceil((marketing_seller_price * sales_percent_fbs) / 100) if marketing_seller_price and sales_percent_fbs else 0
        fbo_transport = math.ceil(float(commissions.get("fbo_direct_flow_trans_max_amount") or 0))
        fbs_transport = math.ceil(float(commissions.get("fbs_direct_flow_trans_max_amount") or 0))
        fbo_delivery = math.ceil(float(commissions.get("fbo_deliv_to_customer_amount") or 0))
        fbs_delivery = math.ceil(float(commissions.get("fbs_deliv_to_customer_amount") or 0))
        dt_value = math.ceil(acquiring + dr_value + fbo_transport + fbo_delivery)
        du_value = math.ceil(acquiring + fbs_transport + fbs_delivery + ds_value)
        auto_action = "🔥" if price_data.get("auto_action_enabled") else "🔕"

        action_titles = []
        for a in marketing_actions:
            if not isinstance(a, dict):
                continue
            title = (a.get("title") or "").strip()
            if title and title not in EXCLUDED_ACTIONS:
                action_titles.append(f"[{title}]")
        action_title = " ".join(action_titles)
        actions_count = len(action_titles)

        rows[str(pid)] = {
            "product_id": pid,
            "HP": math.ceil(acquiring),
            "HQ": math.ceil(sales_percent_fbo),
            "HR": dr_value,
            "HS": fbo_transport,
            "HT": fbo_delivery,
            "HU": commissions.get("fbo_return_flow_amount") or 0,
            "HV": math.ceil(sales_percent_fbs),
            "HW": ds_value,
            "HX": fbs_transport,
            "HY": fbs_delivery,
            "HZ": commissions.get("fbs_return_flow_amount") or 0,
            "IA": dt_value,
            "IB": du_value,
            "IC": "",
            "ID": auto_action,
            "IE": price_data.get("old_price"),
            "IF": price_data.get("min_price"),
            "IG": price_data.get("price"),
            "IH": math.ceil(marketing_seller_price),
            "II": price_data.get("marketing_price"),
            "IJ": COLOR_INDEX_MAP.get(price_indexes.get("color_index"), price_indexes.get("color_index")),
            "IK": action_title,
            "IL": actions_count,
            "IM": price_data.get("net_price"),
        }
    upsert_rows(report, rows)


def sync_fbo_dynamic(shop: Shop) -> None:
    report = get_or_create_report(shop, "fbo_dynamic", "Динамика FBO", "Ежедневные остатки FBO")
    ensure_columns(report, [
        ("sku", "SKU", 10, "text"),
        ("date", "Дата", 20, "date"),
        ("available", "Доступно", 30, "number"),
    ])
    client = OzonClient(client_id=shop.client_id, api_key=shop.token)
    monitor = shop.ozon_reports.filter(code="monitor").first()
    if not monitor:
        return
    skus = [str(r.data.get("sku")).strip() for r in monitor.rows.all() if r.data.get("sku")]
    if not skus:
        return
    items = client.analytics_stocks(skus)
    today = datetime.now().date().isoformat()
    rows: Dict[str, Dict] = {}
    for it in items:
        sku = str(it.get("sku") or "").strip()
        if not sku:
            continue
        key = f"{sku}:{today}"
        rows[key] = {
            "sku": sku,
            "date": today,
            "available": it.get("available_stock_count"),
            "sort_key": today,
        }
    upsert_rows(report, rows)


def sync_orders_fbs_list(shop: Shop) -> None:
    report = get_or_create_report(shop, "orders_fbs_list", "Заказы FBS (список)", "Подробный список FBS")
    ensure_columns(report, [
        ("posting_number", "Номер отправления", 10, "text"),
        ("status", "Статус", 20, "text"),
        ("offer_id", "Артикул", 30, "text"),
        ("quantity", "Количество", 40, "number"),
        ("created_at", "Дата создания", 50, "date"),
        ("shipment_date", "Дата отгрузки", 60, "date"),
        ("product_name", "Наименование", 70, "text"),
        ("cluster_to", "Кластер отправки", 80, "text"),
        ("price", "Цена", 90, "number"),
        ("actions", "Акции", 100, "text"),
    ])
    client = OzonClient(client_id=shop.client_id, api_key=shop.token)
    # reuse /v3/posting/fbs/list
    start = (datetime.now(timezone.utc) - timedelta(days=5)).strftime("%Y-%m-%dT%H:%M:%SZ")
    end = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    offset = 0
    limit = 1000
    rows: Dict[str, Dict] = {}
    while True:
        payload = {
            "dir": "DESC",
            "filter": {"since": start, "to": end},
            "limit": limit,
            "offset": offset,
            "with": {"analytics_data": True, "financial_data": True},
        }
        data = client.post("/v3/posting/fbs/list", payload)
        postings = data.get("result", {}).get("postings", [])
        if not postings:
            break
        for p in postings:
            posting_number = p.get("posting_number") or ""
            products = p.get("products", []) or []
            if not products:
                key = posting_number
                rows[key] = {
                    "posting_number": posting_number,
                    "status": p.get("status"),
                    "created_at": p.get("in_process_at"),
                    "shipment_date": p.get("shipment_date"),
                    "sort_key": p.get("in_process_at") or "",
                }
            for prod in products:
                key = f"{posting_number}:{prod.get('offer_id')}"
                fin = p.get("financial_data") or {}
                fin_prod = (fin.get("products") or [{}])[0]
                rows[key] = {
                    "posting_number": posting_number,
                    "status": p.get("status"),
                    "offer_id": prod.get("offer_id"),
                    "quantity": prod.get("quantity"),
                    "created_at": p.get("in_process_at"),
                    "shipment_date": p.get("shipment_date"),
                    "product_name": prod.get("name"),
                    "cluster_to": fin.get("cluster_to"),
                    "price": fin_prod.get("price"),
                    "actions": ", ".join(fin_prod.get("actions") or []) if isinstance(fin_prod.get("actions"), list) else fin_prod.get("actions"),
                    "sort_key": p.get("in_process_at") or "",
                }
        if len(postings) < limit:
            break
        offset += limit
    upsert_rows(report, rows)


def sync_stocks_analytics(shop: Shop) -> None:
    report = get_or_create_report(shop, "stocks_analytics", "Остатки FBO (аналитика)", "Аналитика остатков по SKU")
    ensure_columns(report, [
        ("sku", "SKU", 10, "text"),
        ("available", "Доступно", 20, "number"),
        ("other", "Прочее", 30, "number"),
        ("requested", "Запрошено", 40, "number"),
        ("return_from_customer", "Возврат от клиента", 50, "number"),
        ("return_to_seller", "Возврат продавцу", 60, "number"),
        ("defect", "Брак", 70, "number"),
        ("transit_defect", "Брак в пути", 80, "number"),
        ("transit", "В пути", 90, "number"),
        ("valid", "Годные", 100, "number"),
    ])
    client = OzonClient(client_id=shop.client_id, api_key=shop.token)
    monitor = shop.ozon_reports.filter(code="monitor").first()
    if not monitor:
        return
    skus = [str(r.data.get("sku")).strip() for r in monitor.rows.all() if r.data.get("sku")]
    if not skus:
        return
    items = client.analytics_stocks(skus)
    rows: Dict[str, Dict] = {}
    for it in items:
        sku = str(it.get("sku") or "").strip()
        if not sku:
            continue
        rows[sku] = {
            "sku": sku,
            "available": it.get("available_stock_count"),
            "other": it.get("other_stock_count"),
            "requested": it.get("requested_stock_count"),
            "return_from_customer": it.get("return_from_customer_stock_count"),
            "return_to_seller": it.get("return_to_seller_stock_count"),
            "defect": it.get("stock_defect_stock_count"),
            "transit_defect": it.get("transit_defect_stock_count"),
            "transit": it.get("transit_stock_count"),
            "valid": it.get("valid_stock_count"),
        }
    upsert_rows(report, rows)


def sync_supplies_fbo(shop: Shop) -> None:
    report = get_or_create_report(shop, "supplies_fbo", "Поставки FBO", "Список поставок и статусов")
    ensure_columns(report, [
        ("supply_id", "ID поставки", 10, "number"),
        ("status", "Статус", 20, "text"),
        ("created_at", "Создано", 30, "date"),
        ("updated_at", "Обновлено", 40, "date"),
        ("warehouse", "Склад", 50, "text"),
        ("cluster", "Кластер", 60, "text"),
        ("items_count", "Позиций", 70, "number"),
    ])
    client = OzonClient(client_id=shop.client_id, api_key=shop.token)
    states = [
        "DATA_FILLING",
        "READY_TO_SUPPLY",
        "ACCEPTED_AT_SUPPLY_WAREHOUSE",
        "IN_TRANSIT",
        "ACCEPTANCE_AT_STORAGE_WAREHOUSE",
        "REPORTS_CONFIRMATION_AWAITING",
    ]
    orders = client.supply_order_list(states=states)
    order_ids = [o.get("supply_order_id") for o in orders if o.get("supply_order_id")]
    if not order_ids:
        return
    rows: Dict[str, Dict] = {}
    for i in range(0, len(order_ids), 100):
        batch_ids = order_ids[i:i + 100]
        data = client.supply_order_get(batch_ids)
        part_orders = []
        if isinstance(data.get("orders"), list):
            part_orders = data["orders"]
        elif isinstance(data.get("result"), dict) and isinstance(data["result"].get("orders"), list):
            part_orders = data["result"]["orders"]
        for o in part_orders:
            sid = o.get("supply_order_id") or o.get("id")
            rows[str(sid)] = {
                "supply_id": sid,
                "status": o.get("status"),
                "created_at": o.get("created_at"),
                "updated_at": o.get("updated_at"),
                "warehouse": (o.get("storage_warehouse") or {}).get("name") if isinstance(o.get("storage_warehouse"), dict) else "",
                "cluster": o.get("cluster_name"),
                "items_count": len(o.get("items") or []),
            }
    upsert_rows(report, rows)


def sync_supply_statuses(shop: Shop) -> None:
    report = get_or_create_report(shop, "supply_statuses", "Движение поставок", "Статусы по SKU")
    statuses = [
        "DATA_FILLING",
        "READY_TO_SUPPLY",
        "ACCEPTED_AT_SUPPLY_WAREHOUSE",
        "IN_TRANSIT",
        "ACCEPTANCE_AT_STORAGE_WAREHOUSE",
        "REPORTS_CONFIRMATION_AWAITING",
    ]
    columns = [("sku", "SKU", 10, "text"), ("total", "Всего", 20, "number")]
    order = 30
    for st in statuses:
        columns.append((f"status_{st.lower()}", st, order, "number"))
        order += 10
        columns.append((f"warehouses_{st.lower()}", f"Склады {st}", order, "text"))
        order += 10
        columns.append((f"clusters_{st.lower()}", f"Кластеры {st}", order, "text"))
        order += 10
    ensure_columns(report, columns)

    client = OzonClient(client_id=shop.client_id, api_key=shop.token)
    supplies = client.supply_order_list(states=statuses)
    warehouses = {w.get("warehouse_id"): w.get("name") for w in client.warehouse_list()}

    rows: Dict[str, Dict] = {}
    for s in supplies:
        status = s.get("status")
        if status not in statuses:
            continue
        supply_id = s.get("supply_order_id") or s.get("id")
        data = client.supply_order_get([supply_id])
        detail = {}
        if isinstance(data.get("orders"), list) and data["orders"]:
            detail = data["orders"][0]
        elif isinstance(data.get("result"), dict) and isinstance(data["result"].get("orders"), list) and data["result"]["orders"]:
            detail = data["result"]["orders"][0]
        items = detail.get("items") or detail.get("products") or []
        warehouse_id = detail.get("warehouse_id") or s.get("warehouse_id")
        wh_name = warehouses.get(warehouse_id) or (detail.get("storage_warehouse") or {}).get("name") or (f"Склад {warehouse_id}" if warehouse_id else "")
        cluster_name = detail.get("cluster_name") or s.get("cluster_name") or ""
        for it in items:
            sku = str(it.get("sku") or "").strip()
            if not sku:
                continue
            qty = int(it.get("quantity") or 0)
            rec = rows.setdefault(sku, {"sku": sku, "total": 0})
            rec["total"] += qty
            rec[f"status_{status.lower()}"] = rec.get(f"status_{status.lower()}", 0) + qty
            wh_key = f"warehouses_{status.lower()}"
            cl_key = f"clusters_{status.lower()}"
            if wh_name:
                rec[wh_key] = (rec.get(wh_key, "") + f"{wh_name}: {qty}\n").strip()
            if cluster_name:
                rec[cl_key] = (rec.get(cl_key, "") + f"🔹 {cluster_name}\n").strip()
    upsert_rows(report, rows)


def sync_stocks_by_cluster(shop: Shop) -> None:
    report = get_or_create_report(shop, "stocks_by_cluster", "Остатки по кластерам", "Остатки SKU по каждому кластеру")
    ensure_columns(report, [
        ("sku", "SKU", 10, "text"),
        ("cluster_id", "Кластер ID", 20, "number"),
        ("available", "Доступно", 30, "number"),
        ("other", "Прочее", 40, "number"),
        ("requested", "Запрошено", 50, "number"),
        ("return_from_customer", "Возврат от клиента", 60, "number"),
        ("return_to_seller", "Возврат продавцу", 70, "number"),
        ("defect", "Брак", 80, "number"),
        ("transit_defect", "Брак в пути", 90, "number"),
        ("transit", "В пути", 100, "number"),
        ("valid", "Годные", 110, "number"),
        ("ads", "ADS", 120, "number"),
        ("days_without_sales", "Дней без продаж", 130, "number"),
        ("idc", "IDC", 140, "number"),
        ("turnover_grade", "Ликвидность", 150, "text"),
    ])
    client = OzonClient(client_id=shop.client_id, api_key=shop.token)
    monitor = shop.ozon_reports.filter(code="monitor").first()
    if not monitor:
        return
    skus = [str(r.data.get("sku")).strip() for r in monitor.rows.all() if r.data.get("sku")]
    if not skus:
        return
    clusters = client.cluster_list()
    rows: Dict[str, Dict] = {}
    for c in clusters:
        cid = c.get("cluster_id") or c.get("id")
        if not cid:
            continue
        items = client.analytics_stocks(skus, cluster_ids=[cid])
        for it in items:
            sku = str(it.get("sku") or "").strip()
            if not sku:
                continue
            key = f"{sku}:{cid}"
            rows[key] = {
                "sku": sku,
                "cluster_id": cid,
                "available": it.get("available_stock_count"),
                "other": it.get("other_stock_count"),
                "requested": it.get("requested_stock_count"),
                "return_from_customer": it.get("return_from_customer_stock_count"),
                "return_to_seller": it.get("return_to_seller_stock_count"),
                "defect": it.get("stock_defect_stock_count"),
                "transit_defect": it.get("transit_defect_stock_count"),
                "transit": it.get("transit_stock_count"),
                "valid": it.get("valid_stock_count"),
                "ads": it.get("ads"),
                "days_without_sales": it.get("days_without_sales"),
                "idc": it.get("idc"),
                "turnover_grade": it.get("turnover_grade"),
            }
    upsert_rows(report, rows)


def _get_cluster_slot_map(shop: Shop, codes: List[str]) -> Dict[str, int]:
    slots = {s.code: s.cluster_id for s in shop.ozon_cluster_slots.filter(code__in=codes)}
    if len(slots) == len(codes):
        return slots
    client = OzonClient(client_id=shop.client_id, api_key=shop.token)
    clusters = client.cluster_list("CLUSTER_TYPE_OZON") + client.cluster_list("CLUSTER_TYPE_CIS")
    ids = [c.get("id") or c.get("cluster_id") for c in clusters if c.get("id") or c.get("cluster_id")]
    ids = [int(x) for x in ids if x is not None]
    fill_ids = ids[: len(codes)]
    for code, cid in zip(codes, fill_ids):
        slots.setdefault(code, cid)
        OzonClusterSlot.objects.update_or_create(shop=shop, code=code, defaults={"cluster_id": cid})
    return slots


def sync_stocks_analytics_full(shop: Shop) -> None:
    report = get_or_create_report(shop, "stocks_analytics_full", "Остатки FBO (полная)", "Полная таблица AC..DI")
    columns = []
    col_order = 10
    for key, label, dtype in [
        ("AC", "AC", "number"),
        ("AD", "AD", "number"),
        ("AE", "AE", "number"),
        ("AF", "AF", "number"),
        ("AG", "AG", "number"),
        ("AH", "AH", "number"),
        ("AI", "AI", "number"),
        ("AJ", "AJ", "number"),
        ("AK", "AK", "number"),
        ("AL", "AL", "text"),
    ]:
        columns.append((key, label, col_order, dtype))
        col_order += 10
    for code in [
        "AM","AN","AO","AP","AQ","AR","AS","AT","AU","AV","AW","AX","AY","AZ","BA","BB","BC","BD","BE","BF"
    ]:
        columns.append((code, code, col_order, "number"))
        col_order += 10
    for code in [
        "BG","BH","BI","BJ","BK","BL","BM","BN","BO","BP","BQ","BR","BS","BT","BU","BV","BW","BX","BY","BZ","CA","CB","CC"
    ]:
        columns.append((code, code, col_order, "text"))
        col_order += 10
    for code, dtype in [
        ("CD","number"), ("CE","number"), ("CF","number"), ("CG","text"),
        ("CH","number"), ("CI","number"), ("CJ","number"), ("CK","text"),
        ("CL","number"), ("CM","number"), ("CN","number"), ("CO","text"),
        ("CP","number"), ("CQ","number"), ("CR","number"), ("CS","text"),
        ("CT","number"), ("CU","number"), ("CV","number"), ("CW","text"),
        ("CX","number"), ("CY","number"), ("CZ","number"), ("DA","text"),
        ("DB","number"), ("DC","number"), ("DD","number"), ("DE","text"),
        ("DF","number"), ("DG","number"), ("DH","number"), ("DI","text"),
    ]:
        columns.append((code, code, col_order, dtype))
        col_order += 10
    ensure_columns(report, columns)

    monitor = shop.ozon_reports.filter(code="monitor").first()
    if not monitor:
        return
    skus = [str(r.data.get("sku")).strip() for r in monitor.rows.all() if r.data.get("sku")]
    if not skus:
        return
    client = OzonClient(client_id=shop.client_id, api_key=shop.token)

    items_all = client.analytics_stocks(skus)
    sums_all: Dict[str, Dict[str, object]] = {}
    for it in items_all:
        sku = str(it.get("sku") or "").strip()
        if not sku:
            continue
        available = it.get("available_stock_count") or 0
        sums_all[sku] = {
            "AC": available if available != 0 else 0,
            "AD": it.get("other_stock_count") or "",
            "AE": it.get("requested_stock_count") or "",
            "AF": it.get("return_from_customer_stock_count") or "",
            "AG": it.get("return_to_seller_stock_count") or "",
            "AH": it.get("stock_defect_stock_count") or "",
            "AI": it.get("transit_defect_stock_count") or "",
            "AJ": it.get("transit_stock_count") or "",
            "AK": it.get("valid_stock_count") or "",
            "AL": "",
        }
        sums_all[sku]["CD"] = _round_two_half_up(it.get("ads")) or ""
        sums_all[sku]["CE"] = it.get("days_without_sales") if it.get("days_without_sales") is not None else ""
        sums_all[sku]["CF"] = _round_two_half_up(it.get("idc")) if it.get("idc") is not None else ""
        sums_all[sku]["CG"] = _normalize_turnover(it.get("turnover_grade"))

    cluster_cols = [
        "AM","AN","AO","AP","AQ","AR","AS","AT","AU","AV","AW","AX","AY","AZ","BA","BB","BC","BD","BE","BF"
    ]
    slot_map = _get_cluster_slot_map(shop, cluster_cols)
    cluster_items: Dict[str, Dict[str, int]] = {c: {} for c in cluster_cols}
    for col, cid in slot_map.items():
        items = client.analytics_stocks(skus, cluster_ids=[cid])
        for it in items:
            sku = str(it.get("sku") or "").strip()
            if not sku:
                continue
            cluster_items[col][sku] = it.get("available_stock_count") or 0

    single_slots = ["CK2","CO2","CS2","CW2","DA2","DE2","DI2"]
    single_cols = [
        ("CH","CI","CJ","CK"),
        ("CL","CM","CN","CO"),
        ("CP","CQ","CR","CS"),
        ("CT","CU","CV","CW"),
        ("CX","CY","CZ","DA"),
        ("DB","DC","DD","DE"),
        ("DF","DG","DH","DI"),
    ]
    single_map = _get_cluster_slot_map(shop, single_slots)
    single_by_sku: Dict[str, Dict[str, Dict[str, object]]] = {}
    for code, cid in single_map.items():
        items = client.analytics_stocks(skus, cluster_ids=[cid])
        for it in items:
            sku = str(it.get("sku") or "").strip()
            if not sku:
                continue
            single_by_sku.setdefault(sku, {})[code] = {
                "ads": _round_two_half_up(it.get("ads")) or "",
                "days_without_sales": it.get("days_without_sales") if it.get("days_without_sales") is not None else "",
                "idc": _round_two_half_up(it.get("idc")) if it.get("idc") is not None else "",
                "turnover_grade": _normalize_turnover(it.get("turnover_grade")),
            }

    rows: Dict[str, Dict] = {}
    for sku in skus:
        data = sums_all.get(sku, {}).copy()
        for col in cluster_cols:
            v = cluster_items.get(col, {}).get(sku, 0)
            data[col] = v if v != 0 else ""
        for code in [
            "BG","BH","BI","BJ","BK","BL","BM","BN","BO","BP","BQ","BR","BS","BT","BU","BV","BW","BX","BY","BZ","CA","CB","CC"
        ]:
            data[code] = ""
        for (code, cols) in zip(single_slots, single_cols):
            block = single_by_sku.get(sku, {}).get(code, {})
            data[cols[0]] = block.get("ads", "")
            data[cols[1]] = block.get("days_without_sales", "")
            data[cols[2]] = block.get("idc", "")
            data[cols[3]] = block.get("turnover_grade", "")
        rows[sku] = data
    upsert_rows(report, rows)


def sync_supply_statuses_full(shop: Shop) -> None:
    report = get_or_create_report(shop, "supply_statuses_full", "Статусы поставок (полная)", "BW..CB как в таблице")
    status_cols = {
        "DATA_FILLING": "BW",
        "READY_TO_SUPPLY": "BX",
        "ACCEPTED_AT_SUPPLY_WAREHOUSE": "BY",
        "IN_TRANSIT": "BZ",
        "ACCEPTANCE_AT_STORAGE_WAREHOUSE": "CA",
        "REPORTS_CONFIRMATION_AWAITING": "CB",
    }
    columns = [("sku", "SKU", 10, "text"), ("BP", "BP всего", 20, "number")]
    order = 30
    for st, col in status_cols.items():
        columns.append((col, col, order, "text"))
        order += 10
    ensure_columns(report, columns)

    monitor = shop.ozon_reports.filter(code="monitor").first()
    if not monitor:
        return
    skus = [str(r.data.get("sku")).strip() for r in monitor.rows.all() if r.data.get("sku")]
    if not skus:
        return
    client = OzonClient(client_id=shop.client_id, api_key=shop.token)
    states = list(status_cols.keys())
    supplies = client.supply_order_list(states=states)
    order_ids = [o.get("supply_order_id") for o in supplies if o.get("supply_order_id")]
    if not order_ids:
        return

    # map warehouse -> cluster
    cluster_map: Dict[str, str] = {}
    for cl in client.cluster_list("CLUSTER_TYPE_OZON") + client.cluster_list("CLUSTER_TYPE_CIS"):
        cname = cl.get("name") or cl.get("cluster_name")
        for wh in cl.get("warehouses") or []:
            wid = str(wh.get("warehouse_id") or wh.get("id") or "")
            if wid and cname:
                cluster_map[wid] = cname
    warehouses = {str(w.get("warehouse_id")): w.get("name") for w in client.warehouse_list() if w.get("warehouse_id")}

    rows: Dict[str, Dict] = {sku: {"sku": sku, "BP": 0, "sort_key": "1"} for sku in skus}
    column_sums = {status_cols[st]: 0 for st in status_cols}
    for i in range(0, len(order_ids), 100):
        batch = order_ids[i:i + 100]
        data = client.supply_order_get(batch)
        orders = data.get("orders") or (data.get("result") or {}).get("orders") or []
        for o in orders:
            status = o.get("status")
            if status not in status_cols:
                continue
            col = status_cols[status]
            for it in o.get("items") or []:
                sku = str(it.get("sku") or "").strip()
                if sku not in rows:
                    continue
                qty = int(it.get("quantity") or 0)
                rows[sku]["BP"] += qty
                column_sums[col] += qty
                wh_id = str(o.get("storage_warehouse_id") or "") or str((o.get("storage_warehouse") or {}).get("warehouse_id") or "")
                wh_name_raw = warehouses.get(wh_id) or (o.get("storage_warehouse") or {}).get("name") or f"Склад {wh_id}" if wh_id else ""
                wh_name = _normalize_display_name(wh_name_raw)
                cl_name_raw = cluster_map.get(wh_id, "")
                cl_name = _normalize_display_name(cl_name_raw) if cl_name_raw else ""
                cell_val = rows[sku].get(col, "")
                line = f"{wh_name}: {qty}"
                cluster_lines = []
                if cl_name:
                    cluster_lines.append(f"🔹 {cl_name}")
                cluster_block = "\n".join(cluster_lines)
                combined = line if not cluster_block else f"{line}\n{cluster_block}"
                rows[sku][col] = (cell_val + "\n" + combined).strip() if cell_val else combined

    meta = {"sku": "Итого", "BP": "", "sort_key": "0"}
    for col, total in column_sums.items():
        meta[col] = total
    rows["__column_sums__"] = meta
    upsert_rows(report, rows)
