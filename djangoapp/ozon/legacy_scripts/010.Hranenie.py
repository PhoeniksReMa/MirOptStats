import argparse
import json
import os
import re
import sys
import time
import warnings
from datetime import datetime, date, timedelta
from urllib.parse import urlparse

import requests
import openpyxl

# --- suppress noisy openpyxl warning ---
warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
    category=UserWarning,
)

# ===== Google Sheets deps =====
try:
    from googleapiclient.discovery import build
    from google.oauth2.service_account import Credentials as SACredentials
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
except ImportError as e:
    raise ImportError(
        "Не установлены библиотеки Google API. Установите:\n"
        "pip install --upgrade google-api-python-client google-auth google-auth-oauthlib google-auth-httplib2"
    ) from e

SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

# ====== OZON API ======
BASE_URL = "https://api-seller.ozon.ru"
CREATE_ENDPOINT = "/v1/report/placement/by-products/create"
INFO_ENDPOINT = "/v1/report/info"


# ====== Time helpers ======
def get_moscow_today_date() -> date:
    try:
        from zoneinfo import ZoneInfo  # Python 3.9+
        msk = ZoneInfo("Europe/Moscow")
        return datetime.now(msk).date()
    except Exception:
        return date.today()


# ====== Read config ======
def read_config(path: str = "API.txt") -> tuple[str, str, str, str]:
    """
    API.txt без заголовков (4 строки):
    1) Client-ID (OZON)
    2) API-KEY   (OZON)
    3) ID Google таблицы
    4) Лист записи
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"Не найден файл: {path}")

    with open(path, "r", encoding="utf-8") as f:
        lines = [line.strip() for line in f.readlines() if line.strip()]

    if len(lines) < 4:
        raise ValueError(
            "В API.txt должно быть 4 непустые строки:\n"
            "1) Client-Id\n2) Api-Key\n3) ID Google таблицы\n4) Имя листа (вкладки)"
        )

    return lines[0], lines[1], lines[2], lines[3]


# ====== OZON HTTP ======
def ozon_post(client_id: str, api_key: str, endpoint: str, payload: dict) -> dict:
    url = BASE_URL + endpoint
    headers = {
        "Client-Id": client_id,
        "Api-Key": api_key,
        "Content-Type": "application/json",
        "Accept": "application/json",
    }

    try:
        r = requests.post(url, headers=headers, json=payload, timeout=60)
    except requests.RequestException as e:
        raise RuntimeError(f"[OZON] Ошибка сети при запросе {endpoint}: {e}")

    if not r.ok:
        raise RuntimeError(f"[OZON] HTTP {r.status_code} при запросе {endpoint}. Ответ: {r.text}")

    try:
        return r.json()
    except ValueError:
        raise RuntimeError(f"[OZON] Не удалось разобрать JSON от {endpoint}. Текст ответа: {r.text}")


def create_placement_report(client_id: str, api_key: str, date_from: str, date_to: str) -> str:
    payload = {"date_from": date_from, "date_to": date_to}
    data = ozon_post(client_id, api_key, CREATE_ENDPOINT, payload)

    code = data.get("code")
    if not code:
        raise RuntimeError(f"[OZON] В ответе create нет поля 'code'. Ответ: {json.dumps(data, ensure_ascii=False)}")
    return code


def get_report_info(client_id: str, api_key: str, code: str) -> dict:
    payload = {"code": code}
    data = ozon_post(client_id, api_key, INFO_ENDPOINT, payload)

    result = data.get("result")
    if not result:
        raise RuntimeError(f"[OZON] В ответе info нет поля 'result'. Ответ: {json.dumps(data, ensure_ascii=False)}")
    return result


def wait_report_ready(
    client_id: str,
    api_key: str,
    code: str,
    poll_seconds: int = 5,
    max_wait_seconds: int = 600,
) -> dict:
    started = time.time()
    last_status = None

    while True:
        result = get_report_info(client_id, api_key, code)
        status = result.get("status")
        error = result.get("error", "")

        if status != last_status:
            if status == "processing":
                print("[OZON] Отчёт формируется...")
            elif status == "success":
                print("[OZON] Отчёт готов.")
            else:
                print(f"[OZON] Статус: {status}")
            last_status = status

        if status == "success":
            if not result.get("file"):
                raise RuntimeError(
                    f"[OZON] Статус success, но нет ссылки 'file'. Ответ: {json.dumps(result, ensure_ascii=False)}"
                )
            return result

        if status in ("failed", "error"):
            raise RuntimeError(f"[OZON] Отчёт завершился с ошибкой. status={status}, error={error}")

        if time.time() - started > max_wait_seconds:
            raise TimeoutError(
                f"[OZON] Не дождались готовности отчёта за {max_wait_seconds} секунд. Последний status={status}"
            )

        time.sleep(poll_seconds)


def download_file(url: str, out_path: str) -> None:
    try:
        r = requests.get(url, stream=True, timeout=120)
        r.raise_for_status()
    except requests.RequestException as e:
        raise RuntimeError(f"[FILE] Не удалось скачать файл отчёта: {e}")

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    with open(out_path, "wb") as f:
        for chunk in r.iter_content(chunk_size=1024 * 256):
            if chunk:
                f.write(chunk)


def guess_extension_from_url(file_url: str) -> str:
    path = urlparse(file_url).path
    return os.path.splitext(path)[1].lower()


# ====== RAW XLSX aggregation ======
def _to_float(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def _to_int(v) -> int:
    return int(round(_to_float(v)))


def _parse_date(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def _find_date_column(headers: list[str]) -> str | None:
    if "Дата" in headers:
        return "Дата"
    candidates = [h for h in headers if "Дата" in h]
    if not candidates:
        return None
    for h in candidates:
        if "ежеднев" in h.lower():
            return h
    return candidates[0]


def aggregate_from_raw_xlsx(raw_xlsx_path: str, snapshot_date_to: str) -> list[dict]:
    """
    DK = cost_total за период (целое)
    DL = qty_paid (на дату среза)
    DM = forecast_28_days_rub = cost_on_snap_day * 28 (целое)
    DN = wh_count (на дату среза)
    NOTE (только DN) = "🔹 СКЛАД - N шт" построчно
    """
    if not os.path.exists(raw_xlsx_path):
        raise FileNotFoundError(f"[FILE] RAW файл не найден: {raw_xlsx_path}")

    target_date = datetime.fromisoformat(snapshot_date_to).date()

    wb = openpyxl.load_workbook(raw_xlsx_path, data_only=True, read_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    idx = {name: i for i, name in enumerate(headers)}

    date_col = _find_date_column(headers)
    if not date_col:
        raise RuntimeError("[XLSX] Не нашёл колонку с датой в RAW XLSX.")

    required = [
        date_col,
        "Склад",
        "SKU",
        "Кол-во экземпляров",
        "Кол-во платных экземпляров",
        "Начисленная стоимость размещения",
    ]
    missing = [c for c in required if c not in idx]
    if missing:
        raise RuntimeError(
            "[XLSX] Не нашёл ожидаемые колонки в XLSX.\n"
            f"Отсутствуют: {missing}\n"
            f"Найденные заголовки: {headers}"
        )

    cost_sum_by_sku: dict[int, float] = {}
    snap_by_sku_day: dict[int, dict[date, dict]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        d = _parse_date(row[idx[date_col]])
        if d is None:
            continue

        sku_raw = row[idx["SKU"]]
        sku = _to_int(sku_raw) if sku_raw is not None else 0
        if sku == 0:
            continue

        wh_raw = row[idx["Склад"]]
        wh = str(wh_raw).strip() if wh_raw is not None else ""
        if not wh:
            wh = "Без названия"

        qty_total = _to_int(row[idx["Кол-во экземпляров"]])
        qty_paid = _to_int(row[idx["Кол-во платных экземпляров"]])
        cost = _to_float(row[idx["Начисленная стоимость размещения"]])

        cost_sum_by_sku[sku] = cost_sum_by_sku.get(sku, 0.0) + cost

        if sku not in snap_by_sku_day:
            snap_by_sku_day[sku] = {}
        if d not in snap_by_sku_day[sku]:
            snap_by_sku_day[sku][d] = {
                "qty_paid_total": 0,
                "daily_cost_total": 0.0,
                "wh_qty_total": {},
                "warehouses_nonzero": set(),
            }

        s = snap_by_sku_day[sku][d]
        s["qty_paid_total"] += qty_paid
        s["daily_cost_total"] += cost

        s["wh_qty_total"][wh] = s["wh_qty_total"].get(wh, 0) + qty_total
        if qty_total > 0:
            s["warehouses_nonzero"].add(wh)

    all_skus = sorted(set(list(cost_sum_by_sku.keys()) + list(snap_by_sku_day.keys())))

    rows_for_gs: list[dict] = []
    for sku in all_skus:
        cost_total = float(cost_sum_by_sku.get(sku, 0.0))

        day_map = snap_by_sku_day.get(sku, {})
        if day_map:
            available_days = sorted(day_map.keys())
            snap_day = target_date if target_date in day_map else available_days[-1]
            snap = day_map[snap_day]

            dl_qty_paid = int(snap["qty_paid_total"])
            daily_cost = float(snap["daily_cost_total"])

            # DM: прогноз 28 дней (целое)
            forecast_28 = int(round(daily_cost * 28.0))

            wh_qty_total: dict[str, int] = snap["wh_qty_total"]
            wh_nonzero = sorted(set(snap["warehouses_nonzero"]))
            wh_count = len(wh_nonzero)

            note_lines = [f"🔹 {wh} - {int(wh_qty_total.get(wh, 0))} шт" for wh in wh_nonzero]
            note = "\n".join(note_lines)
        else:
            dl_qty_paid = 0
            forecast_28 = 0
            wh_count = 0
            note = ""

        rows_for_gs.append(
            {
                "sku": sku,
                "dk": int(round(cost_total)),   # целое
                "dl": int(dl_qty_paid),
                "dm": int(forecast_28),         # целое
                "dn": int(wh_count),
                "note": note,
            }
        )

    return rows_for_gs


# ===== Google Sheets helpers =====
def _col_to_number(col: str) -> int:
    col = col.strip().upper()
    n = 0
    for ch in col:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Bad column: {col}")
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _norm_sku(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    try:
        return str(int(float(s.replace(",", "."))))
    except Exception:
        pass
    return re.sub(r"\D+", "", s)


def _a1(sheet_name: str, a1_range: str) -> str:
    safe = sheet_name.replace("'", "''")
    return f"'{safe}'!{a1_range}"


def get_sheets_service(credentials_path: str = "credentials.json", token_path: str = "token.json"):
    if not os.path.exists(credentials_path):
        raise FileNotFoundError(f"[GS] Не найден {credentials_path} рядом со скриптом")

    with open(credentials_path, "r", encoding="utf-8") as f:
        cred_json = json.load(f)

    # service account
    if isinstance(cred_json, dict) and cred_json.get("type") == "service_account":
        creds = SACredentials.from_service_account_file(credentials_path, scopes=SCOPES)
        return build("sheets", "v4", credentials=creds)

    # oauth client
    creds = None
    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(credentials_path, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(token_path, "w", encoding="utf-8") as token:
            token.write(creds.to_json())

    return build("sheets", "v4", credentials=creds)


def get_sheet_id_and_rowcount(service, spreadsheet_id: str, sheet_name: str) -> tuple[int, int]:
    meta = service.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets(properties(sheetId,title,gridProperties(rowCount)))",
    ).execute()

    for sh in meta.get("sheets", []):
        props = sh.get("properties", {})
        if props.get("title") == sheet_name:
            sheet_id = int(props.get("sheetId"))
            row_count = int((props.get("gridProperties") or {}).get("rowCount", 0))
            return sheet_id, row_count

    raise RuntimeError(f"[GS] Не найден лист (вкладка) '{sheet_name}' в таблице {spreadsheet_id}")


def _build_row_index_by_sku_in_column_griddata(
    service,
    spreadsheet_id: str,
    sheet_name: str,
    sku_col_letter: str = "K",
    start_row: int = 5,
    max_rows_cap: int = 50000,
) -> dict[str, list[int]]:
    _sheet_id, row_count = get_sheet_id_and_rowcount(service, spreadsheet_id, sheet_name)
    if row_count <= 0:
        return {}

    end_row = min(row_count, max_rows_cap)
    if end_row < start_row:
        return {}

    rng = _a1(sheet_name, f"{sku_col_letter}{start_row}:{sku_col_letter}{end_row}")

    resp = service.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        ranges=[rng],
        includeGridData=True,
        fields="sheets(data(rowData(values(formattedValue))))",
    ).execute()

    sheets = resp.get("sheets", [])
    if not sheets:
        return {}

    data = (sheets[0].get("data") or [])
    if not data:
        return {}

    row_data = (data[0].get("rowData") or [])
    row_map: dict[str, list[int]] = {}

    for i, rd in enumerate(row_data):
        values = rd.get("values") or []
        cell = values[0] if values else {}
        sku_raw = cell.get("formattedValue", "")
        sku = _norm_sku(sku_raw)
        if sku:
            gs_row = start_row + i
            row_map.setdefault(sku, []).append(gs_row)

    return row_map


def clear_values_dk_dn(service, spreadsheet_id: str, sheet_name: str, start_row: int, end_row: int):
    """Очищает значения в DK, DL, DM, DN."""
    if end_row < start_row:
        return
    rng = _a1(sheet_name, f"DK{start_row}:DN{end_row}")
    service.spreadsheets().values().clear(
        spreadsheetId=spreadsheet_id,
        range=rng,
        body={}
    ).execute()


def clear_notes_in_dn(service, spreadsheet_id: str, sheet_id: int, start_row: int, end_row: int):
    """Очищает все старые примечания (notes) только в DN."""
    if end_row < start_row:
        return
    dn_col = _col_to_number("DN")
    req = {
        "requests": [
            {
                "repeatCell": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": start_row - 1,
                        "endRowIndex": end_row,
                        "startColumnIndex": dn_col - 1,
                        "endColumnIndex": dn_col,
                    },
                    "cell": {"note": ""},
                    "fields": "note",
                }
            }
        ]
    }
    service.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body=req).execute()


def write_totals_values(service, spreadsheet_id: str, sheet_name: str, total_row: int, totals: list[int]):
    """
    totals = [sum_DK, sum_DL, sum_DM, sum_DN]
    Пишет в DK{total_row}:DN{total_row} числа (НЕ формулы).
    """
    rng = _a1(sheet_name, f"DK{total_row}:DN{total_row}")
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=rng,
        valueInputOption="RAW",
        body={"values": [totals]},
    ).execute()


def push_to_google_sheet(
    rows_for_gs: list[dict],
    spreadsheet_id: str,
    sheet_name: str,
    credentials_path: str = "credentials.json",
    sku_col_letter: str = "K",
    start_row: int = 5,
    totals_row: int = 3,  # итоги в строку 3
):
    """
    Перед записью:
      - очищаем значения DK,DL,DM,DN
      - очищаем примечания DN

    Затем:
      - пишем только где DL>0
      - примечания ставим только в DN

    Итоги:
      - считаем в памяти суммы по фактически записанным строкам (учитывая дубли SKU в колонке K)
      - записываем итоговые значения в DK3..DN3 как числа
    """
    service = get_sheets_service(credentials_path=credentials_path)
    sheet_id, row_count = get_sheet_id_and_rowcount(service, spreadsheet_id, sheet_name)

    print(f"[GS] Чистим диапазон DK{start_row}:DN{row_count} и примечания DN...")
    clear_values_dk_dn(service, spreadsheet_id, sheet_name, start_row, row_count)
    clear_notes_in_dn(service, spreadsheet_id, sheet_id, start_row, row_count)
    print(f"[GS] Ок: очищено DK{start_row}:DN{row_count} + notes(DN)")

    row_map = _build_row_index_by_sku_in_column_griddata(
        service,
        spreadsheet_id,
        sheet_name,
        sku_col_letter=sku_col_letter,
        start_row=start_row,
    )

    dn_col_num = _col_to_number("DN")

    value_updates = []
    note_requests = []

    updated_rows = 0
    matched_skus = 0
    skipped_skus = 0
    notes_set = 0

    # SKU, которых нет в Google Sheet (в колонке sku_col_letter)
    not_found_skus: list[str] = []

    # Итоги по фактически обновлённым строкам (аналог SUM(DK5:DK) после очистки, пустые=0)
    totals_dk = 0
    totals_dl = 0
    totals_dm = 0
    totals_dn = 0

    def chunks(lst, n):
        for i in range(0, len(lst), n):
            yield lst[i:i + n]

    for r in rows_for_gs:
        sku_key = _norm_sku(r.get("sku"))
        if not sku_key:
            skipped_skus += 1
            continue

        gs_rows = row_map.get(sku_key)
        if not gs_rows:
            skipped_skus += 1
            not_found_skus.append(sku_key)
            continue

        matched_skus += 1

        dl = int(r.get("dl", 0))
        if dl == 0:
            continue  # после очистки останется пусто

        dk = int(r.get("dk", 0))
        dm = int(r.get("dm", 0))
        dn = int(r.get("dn", 0))
        note = (r.get("note") or "").strip()

        for gs_row in gs_rows:
            rng = _a1(sheet_name, f"DK{gs_row}:DN{gs_row}")
            value_updates.append({"range": rng, "values": [[dk, dl, dm, dn]]})
            updated_rows += 1

            totals_dk += dk
            totals_dl += dl
            totals_dm += dm
            totals_dn += dn

            if note:
                notes_set += 1
                note_requests.append(
                    {
                        "repeatCell": {
                            "range": {
                                "sheetId": sheet_id,
                                "startRowIndex": gs_row - 1,
                                "endRowIndex": gs_row,
                                "startColumnIndex": dn_col_num - 1,
                                "endColumnIndex": dn_col_num,
                            },
                            "cell": {"note": note},
                            "fields": "note",
                        }
                    }
                )

    # values
    for part in chunks(value_updates, 500):
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"valueInputOption": "USER_ENTERED", "data": part},
        ).execute()

    # notes
    for part in chunks(note_requests, 500):
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": part},
        ).execute()

    # Итоги в строку 3 (как числа)
    write_totals_values(
        service,
        spreadsheet_id,
        sheet_name,
        total_row=totals_row,
        totals=[totals_dk, totals_dl, totals_dm, totals_dn],
    )

    # красиво покажем не найденные SKU
    not_found_skus = sorted(set(not_found_skus), key=lambda x: (len(x), x))
    if not_found_skus:
        preview = ", ".join(not_found_skus[:50])
        tail = "" if len(not_found_skus) <= 50 else f" ... (+{len(not_found_skus)-50})"
        not_found_msg = f"Не найдены в Google Sheet (колонка {sku_col_letter}): {preview}{tail}"
    else:
        not_found_msg = "Не найденных SKU нет"

    print(
        f"[GS] Обновлено строк (DK..DN, где DL>0): {updated_rows}\n"
        f"[GS] SKU сопоставлено: {matched_skus}, не найдено: {len(not_found_skus)}\n"
        f"[GS] Примечаний DN установлено: {notes_set}\n"
        f"[GS] Итоги: DK{totals_row}={totals_dk}, DL{totals_row}={totals_dl}, DM{totals_row}={totals_dm}, DN{totals_row}={totals_dn}\n"
        f"[GS] {not_found_msg}"
    )


# ====== CLI / MAIN ======
def main():
    parser = argparse.ArgumentParser(
        description="ONLINE: OZON placement by products -> скачать RAW.xlsx -> агрегировать -> записать в Google Sheets (DK-DN) по SKU из K5:K."
    )
    parser.add_argument("--api-file", default="API.txt",
                        help="API.txt (4 строки: Client-Id, Api-Key, spreadsheet_id, sheet_name)")
    parser.add_argument("--credentials", default="credentials.json", help="credentials.json для Google")

    parser.add_argument("--days", type=int, default=28, help="Сколько суток между date_from и date_to (0..30)")
    parser.add_argument("--date-from", default=None, help="YYYY-MM-DD (если не задано — date_to - days)")
    parser.add_argument("--date-to", default=None, help="YYYY-MM-DD (если не задано — сегодня по Москве)")

    parser.add_argument("--poll", type=int, default=5, help="Интервал опроса /v1/report/info (сек)")
    parser.add_argument("--max-wait", type=int, default=600, help="Макс. ожидание готовности отчёта (сек)")
    parser.add_argument("--code", default=None, help="Если уже есть code — пропустить create и сразу ждать/скачать")

    parser.add_argument("--dir", default=".", help="Папка для сохранения RAW")
    parser.add_argument("--prefix", default=None, help="Префикс имени RAW файла")

    parser.add_argument("--gs-sku-col", default="K", help="Колонка в Google Sheet где SKU (по умолчанию K)")
    parser.add_argument("--gs-start-row", type=int, default=5, help="Строка, с которой начинаются данные (по умолчанию 5)")

    args = parser.parse_args()

    if args.days < 0 or args.days > 30:
        raise ValueError("--days должен быть в диапазоне 0..30 (период <= 31 день включительно).")

    client_id, api_key, spreadsheet_id, sheet_name = read_config(args.api_file)

    # dates
    date_to = args.date_to or get_moscow_today_date().isoformat()
    dt_to = datetime.fromisoformat(date_to).date()

    if args.date_from:
        date_from = args.date_from
    else:
        date_from = (dt_to - timedelta(days=args.days)).isoformat()

    dt_from = datetime.fromisoformat(date_from).date()
    if (dt_to - dt_from).days > 30:
        raise ValueError("Период слишком большой: (date_to - date_from).days должен быть <= 30.")

    print(f"[OZON] Период: {date_from} .. {date_to}")

    # create/wait
    if args.code:
        code = args.code
        print(f"[OZON] Используем code: {code}")
    else:
        code = create_placement_report(client_id, api_key, date_from, date_to)
        print(f"[OZON] Отчёт создан: {code}")

    result = wait_report_ready(client_id, api_key, code, poll_seconds=args.poll, max_wait_seconds=args.max_wait)

    file_url = result["file"]

    # download RAW
    os.makedirs(args.dir, exist_ok=True)
    safe_code = code.replace("/", "_").replace("\\", "_")
    prefix = args.prefix or f"placement_by_products_{date_from}_to_{date_to}"
    ext = guess_extension_from_url(file_url) or ".xlsx"
    raw_path = os.path.join(args.dir, f"{prefix}_RAW_{safe_code}{ext}")

    download_file(file_url, raw_path)
    print(f"[FILE] RAW сохранён: {raw_path}")

    if not raw_path.lower().endswith(".xlsx"):
        raise RuntimeError("[XLSX] Ozon вернул не XLSX — обработка возможна только для .xlsx")

    # aggregate + push
    rows_for_gs = aggregate_from_raw_xlsx(raw_path, snapshot_date_to=date_to)

    print(f"[GS] Запись в таблицу: {spreadsheet_id} / лист: '{sheet_name}'")
    push_to_google_sheet(
        rows_for_gs=rows_for_gs,
        spreadsheet_id=spreadsheet_id,
        sheet_name=sheet_name,
        credentials_path=args.credentials,
        sku_col_letter=args.gs_sku_col,
        start_row=args.gs_start_row,
        totals_row=3,
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"[error] {e}", file=sys.stderr)
        sys.exit(1)
